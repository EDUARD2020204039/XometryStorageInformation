"""
Aplicația FastAPI pentru gestionarea ofertelor Xometry
"""
import os
import asyncio
import sys
from contextvars import ContextVar
from logging.handlers import RotatingFileHandler
from uuid import uuid4
from typing import Dict, List, Optional
import logging
from datetime import datetime
from pydantic import BaseModel
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, BackgroundTasks, Request, Form, HTTPException, Depends, Body
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import String as db_String
from sqlalchemy.orm import Session
from dotenv import load_dotenv
import pandas as pd
from pathlib import Path
import shutil
import re
from migrate_database import migrate_database as run_db_migration
from openpyxl import Workbook
import sentry_sdk
from sentry_sdk.integrations.fastapi import FastApiIntegration
from sentry_sdk.integrations.logging import LoggingIntegration

from xometry.db import init_db, get_db
from xometry.models import Offer, Part, Attachment, Order
from xometry.image_utils import download_and_save_image, cleanup_old_images

# Încarcă variabilele de mediu
load_dotenv()
load_dotenv("/app/data/.env", override=True)

_request_id_ctx: ContextVar[str] = ContextVar("request_id", default="-")


class RequestIdFilter(logging.Filter):
    def filter(self, record: logging.LogRecord) -> bool:
        record.request_id = _request_id_ctx.get("-")
        return True


def _env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        return int(raw)
    except ValueError:
        return default


def _env_float(name: str, default: float) -> float:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        return float(raw)
    except ValueError:
        return default


def configure_error_reporting() -> bool:
    dsn = os.getenv("GLITCHTIP_DSN") or os.getenv("SENTRY_DSN")
    if not dsn:
        return False

    environment = os.getenv("APP_ENV", os.getenv("NODE_ENV", "production"))
    release = os.getenv("APP_RELEASE")

    sentry_sdk.init(
        dsn=dsn,
        integrations=[
            FastApiIntegration(),
            LoggingIntegration(level=logging.INFO, event_level=logging.ERROR),
        ],
        environment=environment,
        release=release,
        traces_sample_rate=_env_float("SENTRY_TRACES_SAMPLE_RATE", 0.0),
        send_default_pii=False,
    )
    return True


def configure_logging() -> None:
    app_root = Path(__file__).resolve().parent
    logs_dir = app_root / "logs"
    logs_dir.mkdir(exist_ok=True)

    level_name = os.getenv("LOG_LEVEL", "INFO").upper()
    level = getattr(logging, level_name, logging.INFO)
    max_bytes = _env_int("LOG_MAX_BYTES", 5 * 1024 * 1024)
    backup_count = _env_int("LOG_BACKUP_COUNT", 5)
    file_level_name = os.getenv("LOG_FILE_LEVEL", level_name).upper()
    file_level = getattr(logging, file_level_name, level)

    fmt = "%(asctime)s %(levelname)s [%(name)s] [req:%(request_id)s] %(message)s"
    formatter = logging.Formatter(fmt)
    request_filter = RequestIdFilter()

    root_logger = logging.getLogger()
    root_logger.setLevel(level)
    root_logger.handlers.clear()

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(level)
    console_handler.setFormatter(formatter)
    console_handler.addFilter(request_filter)
    root_logger.addHandler(console_handler)

    app_file_handler = RotatingFileHandler(
        app_root / "app.log",
        maxBytes=max_bytes,
        backupCount=backup_count,
        encoding="utf-8"
    )
    app_file_handler.setLevel(file_level)
    app_file_handler.setFormatter(formatter)
    app_file_handler.addFilter(request_filter)
    root_logger.addHandler(app_file_handler)

    error_file_handler = RotatingFileHandler(
        logs_dir / "error.log",
        maxBytes=max_bytes,
        backupCount=backup_count,
        encoding="utf-8"
    )
    error_file_handler.setLevel(logging.ERROR)
    error_file_handler.setFormatter(formatter)
    error_file_handler.addFilter(request_filter)
    root_logger.addHandler(error_file_handler)

    # Reduce zgomotul HTTP access logs (de obicei util doar la depanare).
    logging.getLogger("uvicorn.access").setLevel(logging.WARNING)


configure_logging()
logger = logging.getLogger(__name__)
error_reporting_enabled = configure_error_reporting()
if error_reporting_enabled:
    logger.info("GlitchTip/Sentry error reporting enabled")
else:
    logger.info("GlitchTip/Sentry error reporting disabled (DSN missing)")

# Inițializează FastAPI

# --- Pydantic Schemas for Documentation ---
class OrderSchema(BaseModel):
    order_id: str
    part_id: Optional[str] = None
    status: Optional[str] = None
    date: Optional[str] = None
    price: Optional[str] = None
    # Alte câmpuri dinamice
    class Config:
        extra = "allow"

class OrderSyncRequest(BaseModel):
    orders: List[OrderSchema]

class OrderCheckResponse(BaseModel):
    exists: bool
    order_id: Optional[str] = None
    date: Optional[str] = None
    status: Optional[str] = None
    price: Optional[str] = None

# --- WebSocket Manager ---
class ConnectionManager:
    def __init__(self):
        # Maps part_id -> List[WebSocket]
        self.active_connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, websocket: WebSocket, part_id: str):
        await websocket.accept()
        if part_id not in self.active_connections:
            self.active_connections[part_id] = []
        self.active_connections[part_id].append(websocket)
        logger.info(f"WS Connected to {part_id}")

    def disconnect(self, websocket: WebSocket, part_id: str):
        if part_id in self.active_connections:
            if websocket in self.active_connections[part_id]:
                self.active_connections[part_id].remove(websocket)
            if not self.active_connections[part_id]:
                del self.active_connections[part_id]
        logger.info(f"WS Disconnected from {part_id}")

    async def send_personal_message(self, message: str, websocket: WebSocket):
        await websocket.send_text(message)

    async def broadcast_to_part(self, message: dict, part_id: str):
        if part_id in self.active_connections:
            for connection in self.active_connections[part_id]:
                try:
                    await connection.send_json(message)
                except Exception as e:
                    logger.error(f"WS Send Error: {e}")

manager = ConnectionManager()

app = FastAPI(title="Xometry Offer Helper", version="1.0.0")

# Configurare CORS pentru extensia Chrome
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def request_context_middleware(request: Request, call_next):
    request_id = request.headers.get("X-Request-ID") or uuid4().hex[:12]
    token = _request_id_ctx.set(request_id)
    started = datetime.utcnow()
    try:
        response = await call_next(request)
    finally:
        elapsed_ms = (datetime.utcnow() - started).total_seconds() * 1000
        logger.debug(
            "HTTP %s %s completed in %.2fms",
            request.method,
            request.url.path,
            elapsed_ms
        )
        _request_id_ctx.reset(token)

    response.headers["X-Request-ID"] = request_id
    return response


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    logger.warning(
        "HTTPException %s %s -> %s (%s)",
        request.method,
        request.url.path,
        exc.status_code,
        exc.detail
    )
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail, "request_id": _request_id_ctx.get("-")}
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    if error_reporting_enabled:
        sentry_sdk.set_tag("request_id", _request_id_ctx.get("-"))
        sentry_sdk.capture_exception(exc)
    logger.exception("Unhandled exception on %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=500,
        content={"detail": "Eroare internă", "request_id": _request_id_ctx.get("-")}
    )

# Inițializează baza de date
database_url = os.getenv('DATABASE_URL', 'sqlite:///xometry_offers.db')
init_db(database_url)

# Configurare template-uri și fișiere statice
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

def _safe_slug(text: str) -> str:
    s = (text or 'reper').strip()
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    return s[:80] if len(s) > 0 else 'reper'


def _extract_canonical_offer_title(text: Optional[str]) -> Optional[str]:
    if not text:
        return None

    match = re.search(r"\b(J-\d+(?:-\d+)?)\b", text, re.IGNORECASE)
    if match:
        return match.group(1).upper()

    match = re.search(r"\b(RFQ-\d+(?:-\d+)?)\b", text, re.IGNORECASE)
    if match:
        return match.group(1).upper()

    match = re.search(r"\b(HJO-\d+(?:-\d+)?)\b", text, re.IGNORECASE)
    if match:
        return match.group(1).upper()

    return None


def _normalize_offer_title(
    raw_title: Optional[str],
    offer_external_id: Optional[str] = None,
    offer_url: Optional[str] = None,
) -> str:
    canonical = _extract_canonical_offer_title(raw_title) or _extract_canonical_offer_title(offer_url)
    if canonical:
        return canonical

    cleaned = (raw_title or "").strip()
    if cleaned:
        return cleaned

    return str(offer_external_id or "")


def _job_tokens(value: Optional[str]) -> list[str]:
    text = (value or "").upper()
    tokens = []
    for pattern in (r"\bHJO-\d+(?:-\d+)?\b", r"\bJ-\d+(?:-\d+)?\b", r"\bRFQ-\d+(?:-\d+)?\b"):
        tokens.extend(match.group(0) for match in re.finditer(pattern, text))
    roots = []
    for token in tokens:
        parts = token.split("-")
        if len(parts) >= 2:
            roots.append("-".join(parts[:2]))
    return list(dict.fromkeys([*tokens, *roots]))


def _part_ids_from_raw(raw: Optional[str]) -> list[str]:
    if not raw:
        return []
    values = []
    for item in raw.split(","):
        item = item.strip()
        if item:
            values.append(item)
    return list(dict.fromkeys(values))


def _offer_public_url(offer: Offer) -> str:
    return f"/offer/{offer.id}"


def _offer_dosar_payload(offer: Offer, reason: str = "") -> dict:
    return {
        "id": offer.id,
        "offer_id": offer.offer_id,
        "title": _normalize_offer_title(offer.title, offer.offer_id, offer.url),
        "url": offer.url,
        "backend_url": _offer_public_url(offer),
        "dosar_id": offer.dosar_id,
        "dosar_path": offer.dosar_path,
        "dosar_allocated": offer.dosar_allocated.isoformat() if offer.dosar_allocated else None,
        "has_dosar": bool(offer.dosar_id),
        "reason": reason,
    }


def _find_offer_by_external_id(db: Session, external_offer_id: str) -> Offer | None:
    offer = db.query(Offer).filter(Offer.offer_id == str(external_offer_id)).first()
    if offer:
        return offer
    if str(external_offer_id).isdigit():
        return db.query(Offer).filter(Offer.id == int(external_offer_id)).first()
    return None


def _find_dosar_references(
    db: Session,
    current_offer: Offer | None,
    external_offer_id: str,
    job_id: Optional[str],
    part_ids: list[str],
) -> list[dict]:
    current_internal_id = current_offer.id if current_offer else None
    references: dict[int, dict] = {}
    tokens = _job_tokens(job_id)

    for offer in db.query(Offer).order_by(Offer.created_at.desc()).limit(1000).all():
        if current_internal_id and offer.id == current_internal_id:
            continue
        haystack = " ".join([
            str(offer.offer_id or ""),
            str(offer.title or ""),
            str(offer.url or ""),
        ]).upper()
        if any(token and token in haystack for token in tokens):
            references[offer.id] = _offer_dosar_payload(offer, "job")

    for part_id in part_ids:
        query = db.query(Part).filter(Part.part_id == part_id)
        for part in query.all():
            offer = db.query(Offer).filter(Offer.id == part.offer_id).first()
            if not offer or (current_internal_id and offer.id == current_internal_id):
                continue
            references[offer.id] = _offer_dosar_payload(offer, f"part:{part_id}")

    return list(references.values())

def _find_deviz_template() -> Optional[Path]:
    root = Path(__file__).resolve().parent
    candidates = [
        root / 'deviz.xlsx',
        root / 'static' / 'deviz.xlsx',
        root / 'static' / 'templates' / 'deviz.xlsx',
        root.parent / 'deviz.xlsx',
    ]
    for c in candidates:
        if c.exists():
            return c
    return None

@app.get("/images/parts/{offer_id}/{filename}")
async def serve_part_image(offer_id: str, filename: str):
    """Servește imaginile reperelor"""
    image_path = f"static/images/parts/{offer_id}/{filename}"
    if os.path.exists(image_path):
        return FileResponse(image_path)
    else:
        raise HTTPException(status_code=404, detail="Imaginea nu a fost găsită")

@app.post("/api/migrate-images")
async def migrate_existing_images(db: Session = Depends(get_db)):
    """Migrează imaginile existente de la URL-uri externe la stocare locală"""
    try:
        # Găsește toate reperele cu imagini externe dar fără imagini locale
        parts_with_external_images = db.query(Part).filter(
            Part.image_url.isnot(None),
            Part.image_url != '',
            Part.local_image_path.is_(None)
        ).all()
        
        migrated_count = 0
        failed_count = 0
        
        for part in parts_with_external_images:
            try:
                # Obține offer_id pentru a crea directorul corect
                offer = db.query(Offer).filter(Offer.id == part.offer_id).first()
                if not offer:
                    continue
                    
                # Încearcă să descarce imaginea
                local_path = download_and_save_image(
                    part.image_url,
                    str(offer.offer_id),
                    part.part_id
                )
                
                if local_path:
                    part.local_image_path = local_path
                    migrated_count += 1
                    logger.info(f"Imagine migrată pentru reperul {part.part_id}")
                else:
                    failed_count += 1
                    logger.warning(f"Nu s-a putut migra imaginea pentru reperul {part.part_id}")
                    
            except Exception as e:
                logger.error(f"Eroare la migrarea imaginii pentru reperul {part.part_id}: {e}")
                failed_count += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Migrare completă! {migrated_count} imagini migrate, {failed_count} eșecuri.",
            "migrated_count": migrated_count,
            "failed_count": failed_count
        }
        
    except Exception as e:
        logger.error(f"Eroare la migrarea imaginilor: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Eroare la migrarea imaginilor: {str(e)}")

@app.post("/api/download-docs")
async def download_docs(request: dict, db: Session = Depends(get_db)):
    """Descarcă documentația pentru o ofertă"""
    try:
        offer_id = request.get("offer_id")
        download_url = request.get("download_url")
        file_name = request.get("file_name")
        page_type = request.get("page_type", "job")
        
        if not all([offer_id, download_url, file_name]):
            return {"success": False, "error": "Parametri lipsă"}
        
        # Creează directorul pentru documentație
        docs_dir = f"static/docs/{offer_id}"
        os.makedirs(docs_dir, exist_ok=True)
        
        # Descarcă fișierul
        import requests
        response = requests.get(download_url, timeout=30)
        response.raise_for_status()
        
        # Salvează fișierul
        file_path = os.path.join(docs_dir, file_name)
        with open(file_path, 'wb') as f:
            f.write(response.content)
        
        logger.info(f"✅ Documentație descărcată: {file_path}")
        
        # Actualizează calea documentației în baza de date
        offer = db.query(Offer).filter(Offer.offer_id == offer_id).first()
        if offer:
            offer.documentation_path = f"/static/docs/{offer_id}/{file_name}"
            db.commit()
            logger.info(f"✅ Calea documentației actualizată în BD: {offer.documentation_path}")
        
        return {
            "success": True,
            "file_path": f"/static/docs/{offer_id}/{file_name}",
            "message": f"Documentația {file_name} descărcată cu succes"
        }
        
    except Exception as e:
        logger.error(f"❌ Eroare la descărcarea documentației: {e}")
        return {"success": False, "error": str(e)}

@app.on_event("startup")
async def startup_event():
    """Eveniment la pornirea aplicației"""
    logger.info("Aplicația Xometry Offer Helper a pornit")
    # Rulează migrarea bazei de date pentru a garanta coloanele noi
    try:
        run_db_migration()
        logger.info("Migrarea bazei de date finalizată")
    except Exception as e:
        logger.warning(f"Nu s-a putut rula migrarea BD: {e}")

@app.get("/", response_class=HTMLResponse)
async def index(request: Request, db: Session = Depends(get_db)):
    """Pagina principală cu lista ofertelor"""
    try:
        # Obține toate ofertele
        offers = db.query(Offer).order_by(Offer.created_at.desc()).all()
        for offer in offers:
            offer.normalized_title = _normalize_offer_title(offer.title, offer.offer_id, offer.url)
        
        return templates.TemplateResponse("index.html", {
            "request": request,
            "offers": offers
        })
    except Exception as e:
        logger.error(f"Eroare la încărcarea paginii principale: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/orders", response_class=HTMLResponse)
async def list_orders(request: Request, db: Session = Depends(get_db)):
    """Afișează dashboard-ul cu istoricul comenzilor"""
    try:
        orders = db.query(Order).order_by(Order.created_at.desc()).all()
        return templates.TemplateResponse("orders.html", {"request": request, "orders": orders})
    except Exception as e:
        logger.error(f"Error listing orders: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/documentation", response_class=HTMLResponse)
async def documentation(request: Request):
    """Pagina de documentație"""
    try:
        return templates.TemplateResponse("documentation.html", {
            "request": request
        })
    except Exception as e:
        logger.error(f"Eroare la încărcarea documentației: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/help", response_class=HTMLResponse)
async def help_page(request: Request):
    """Pagina de ajutor rapid"""
    try:
        return templates.TemplateResponse("help.html", {
            "request": request
        })
    except Exception as e:
        logger.error(f"Eroare la încărcarea paginii de ajutor: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/api/health")
async def health_check():
    """Health check endpoint pentru monitorizare"""
    try:
        return {
            "status": "healthy",
            "timestamp": datetime.utcnow().isoformat(),
            "version": "1.0.0",
            "database": "connected"
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return {
            "status": "unhealthy",
            "timestamp": datetime.utcnow().isoformat(),
            "version": "1.0.0",
            "error": str(e)
        }

@app.post("/api/force-download-docs/{offer_id}")
async def force_download_docs(offer_id: str, db: Session = Depends(get_db)):
    """Forțează descărcarea documentației pentru o ofertă"""
    try:
        # Găsește oferta
        offer = db.query(Offer).filter(Offer.offer_id == offer_id).first()
        if not offer:
            raise HTTPException(status_code=404, detail="Oferta nu a fost găsită")
        
        # Verifică dacă documentația există deja
        if offer.documentation_path and os.path.exists(f"static{offer.documentation_path}"):
            return {
                "success": True,
                "message": "Documentația există deja",
                "file_path": offer.documentation_path
            }
        
        # Încearcă să găsească documentația în directorul static
        docs_dir = f"static/docs/{offer_id}"
        if os.path.exists(docs_dir):
            files = os.listdir(docs_dir)
            if files:
                # Găsește primul fișier
                file_name = files[0]
                offer.documentation_path = f"/static/docs/{offer_id}/{file_name}"
                db.commit()
                
                return {
                    "success": True,
                    "message": "Documentația găsită local",
                    "file_path": offer.documentation_path
                }
        
        return {
            "success": False,
            "message": "Documentația nu a fost găsită. Folosește extensia Chrome pentru a o descărca.",
            "file_path": None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Eroare la forțarea descărcării documentației: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/offer/{offer_id}", response_class=HTMLResponse)
async def offer_detail(request: Request, offer_id: int, db: Session = Depends(get_db)):
    """Pagina de detalii pentru o ofertă"""
    try:
        # Obține oferta
        offer = db.query(Offer).filter(Offer.id == offer_id).first()
        if not offer:
            raise HTTPException(status_code=404, detail="Oferta nu a fost găsită")
        offer.normalized_title = _normalize_offer_title(offer.title, offer.offer_id, offer.url)
            
        # Obține reperele
        parts = db.query(Part).filter(Part.offer_id == offer_id).all()
        
        # Verifică și curăță datele înainte de a le trimite la template
        for part in parts:
            # Asigură-te că toate câmpurile numerice au valori valide
            if part.unit_price is None:
                part.unit_price = 0.0
            if part.discount is None:
                part.discount = 0.0
            if part.quantity is None:
                part.quantity = 1
            if part.total_price is None:
                part.total_price = 0.0
            if part.lead_time is None:
                part.lead_time = 0
        
        return templates.TemplateResponse("offer_detail.html", {
            "request": request,
            "offer": offer,
            "parts": parts
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Eroare la încărcarea detaliilor ofertei: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.post("/update_part")
async def update_part(
    part_id: int = Form(...),
    unit_price: Optional[float] = Form(None),
    discount: Optional[float] = Form(0.0),
    lead_time: Optional[int] = Form(None),
    db: Session = Depends(get_db)
):
    """Actualizează datele unui reper"""
    try:
        # Găsește reperul
        part = db.query(Part).filter(Part.id == part_id).first()
        if not part:
            raise HTTPException(status_code=404, detail="Reperul nu a fost găsit")
            
        # Actualizează datele
        if unit_price is not None:
            part.unit_price = unit_price
        if discount is not None:
            part.discount = discount
        if lead_time is not None:
            part.lead_time = lead_time

            
        # Calculează prețul total
        if part.unit_price and part.quantity:
            part.total_price = part.unit_price * part.quantity * (1 - part.discount / 100)
            
        db.commit()
        
        return {"success": True, "message": "Reperul a fost actualizat cu succes"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Eroare la actualizarea reperului: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.post("/api/part/{part_id}/deviz/create")
async def create_deviz_for_part(part_id: int, db: Session = Depends(get_db)):
    """Creează o copie a fișierului deviz.xlsx pentru reper și returnează linkul."""
    try:
        logger.info(f"[DEVIZ] Cerere creare deviz pentru part_id=%s", part_id)
        part = db.query(Part).filter(Part.id == part_id).first()
        if not part:
            raise HTTPException(status_code=404, detail="Reperul nu a fost găsit")

        # Dacă există deja, returnează linkul existent
        if part.deviz_path and os.path.exists(os.path.join('static', part.deviz_path)):
            return {
                'success': True,
                'message': 'Deviz existent',
                'url': f"/static/{part.deviz_path}",
                'deviz_path': part.deviz_path
            }

        template = _find_deviz_template()
        logger.info(f"[DEVIZ] Șablon găsit: %s", str(template) if template else "None")

        # Obține oferta pentru folderul dedicat
        offer = db.query(Offer).filter(Offer.id == part.offer_id).first()
        offer_folder = str(offer.offer_id) if offer else 'unknown_offer'

        out_dir = Path('static') / 'devize' / offer_folder
        out_dir.mkdir(parents=True, exist_ok=True)

        safe_name = _safe_slug(part.name or part.part_id or 'reper')
        filename = f"deviz_{safe_name}.xlsx"
        dest = out_dir / filename

        # Evită suprascrierea: dacă există, adaugă un sufix numeric
        if dest.exists():
            idx = 2
            while True:
                alt = out_dir / f"deviz_{safe_name}_{idx}.xlsx"
                if not alt.exists():
                    dest = alt
                    break
                idx += 1

        logger.info(f"[DEVIZ] Destinație deviz: %s", str(dest))

        if template and template.exists():
            shutil.copyfile(template, dest)
            logger.info("[DEVIZ] Copiat din șablon")
        else:
            # Fallback: creează un fișier XLSX gol pre-populat minimal
            wb = Workbook()
            ws = wb.active
            ws.title = "Deviz"
            # Headere simple
            ws["A1"] = "Reper"
            ws["B1"] = part.name or part.part_id or "Reper"
            ws["A2"] = "Material"
            ws["B2"] = part.material or "-"
            ws["A3"] = "Dimensiuni (L x W x H mm)"
            l = part.length or 0; w = part.width or 0; h = part.height or 0
            ws["B3"] = f"{l} x {w} x {h}"
            ws["A4"] = "Cantitate"
            ws["B4"] = part.quantity or 1
            ws["A6"] = "Pozitie"
            ws["B6"] = "Descriere"
            ws["C6"] = "UM"
            ws["D6"] = "Cant"
            ws["E6"] = "PU"
            ws["F6"] = "Valoare"
            wb.save(dest)
            logger.info("[DEVIZ] Creat XLSX fallback cu openpyxl")

        # Salvează calea relativă în BD
        rel_path = str(dest.relative_to('static'))
        part.deviz_path = rel_path
        try:
            db.commit()
        except Exception as e:
            logger.warning(f"[DEVIZ] Commit a eșuat, încerc migrarea BD: {e}")
            try:
                run_db_migration()
                db.commit()
            except Exception as e2:
                logger.exception(f"[DEVIZ] Commit eșuat după migrare: {e2}")
                raise

        return {
            'success': True,
            'message': 'Deviz creat',
            'url': f"/static/{rel_path}",
            'deviz_path': rel_path
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Eroare la crearea devizului: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Eroare internă: {str(e)}")

@app.get("/export/{offer_id}/csv")
async def export_csv(offer_id: int, db: Session = Depends(get_db)):
    """Exportă oferta în format CSV"""
    try:
        # Obține oferta și reperele
        offer = db.query(Offer).filter(Offer.id == offer_id).first()
        if not offer:
            raise HTTPException(status_code=404, detail="Oferta nu a fost găsită")
            
        parts = db.query(Part).filter(Part.offer_id == offer_id).all()
        
        # Creează DataFrame-ul
        data = []
        for part in parts:
            data.append({
                'ID Reper': part.part_id,
                'Denumire': part.name,
                'Material': part.material or '',
                'Observații': part.remarks or '',
                'Greutate (kg)': part.weight or '',
                'Lungime (mm)': part.length or '',
                'Lățime (mm)': part.width or '',
                'Înălțime (mm)': part.height or '',
                'Cantitate': part.quantity,
                'Preț unitar (€)': part.unit_price or '',
                'Discount (%)': part.discount or 0,
                'Lead time (zile)': part.lead_time or '',
                'Procese': ', '.join(part.processes) if part.processes else '',
                'Preț total (€)': part.total_price or ''
            })
            
        df = pd.DataFrame(data)
        
        # Salvează CSV-ul
        csv_path = f"exports/offer_{offer_id}.csv"
        os.makedirs("exports", exist_ok=True)
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        return FileResponse(
            csv_path,
            media_type='text/csv',
            filename=f"oferta_{offer.offer_id}.csv"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Eroare la exportul CSV: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/export/{offer_id}/xlsx")
async def export_xlsx(offer_id: int, db: Session = Depends(get_db)):
    """Exportă oferta în format XLSX"""
    try:
        # Obține oferta și reperele
        offer = db.query(Offer).filter(Offer.id == offer_id).first()
        if not offer:
            raise HTTPException(status_code=404, detail="Oferta nu a fost găsită")
            
        parts = db.query(Part).filter(Part.offer_id == offer_id).all()
        
        # Creează DataFrame-ul
        data = []
        for part in parts:
            data.append({
                'ID Reper': part.part_id,
                'Denumire': part.name,
                'Material': part.material or '',
                'Observații': part.remarks or '',
                'Greutate (kg)': part.weight or '',
                'Lungime (mm)': part.width or '',
                'Lățime (mm)': part.width or '',
                'Înălțime (mm)': part.height or '',
                'Cantitate': part.quantity,
                'Preț unitar (€)': part.unit_price or '',
                'Discount (%)': part.discount or 0,
                'Lead time (zile)': part.lead_time or '',
                'Preț total (€)': part.total_price or ''
            })
            
        df = pd.DataFrame(data)
        
        # Salvează XLSX-ul
        xlsx_path = f"exports/offer_{offer_id}.xlsx"
        os.makedirs("exports", exist_ok=True)
        
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Ofertă', index=False)
            
            # Adaugă informații despre ofertă
            offer_info = pd.DataFrame([{
                'Câmp': 'ID Ofertă',
                'Valoare': offer.offer_id
            }, {
                'Câmp': 'Titlu',
                'Valoare': offer.title or ''
            }, {
                'Câmp': 'Client',
                'Valoare': offer.customer or ''
            }, {
                'Câmp': 'URL',
                'Valoare': offer.url
            }, {
                'Câmp': 'Data creării',
                'Valoare': offer.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }])
            
            offer_info.to_excel(writer, sheet_name='Informații Ofertă', index=False)
        
        return FileResponse(
            xlsx_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"oferta_{offer.offer_id}.xlsx"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Eroare la exportul XLSX: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/api/offers")
async def get_offers(db: Session = Depends(get_db)):
    """API endpoint pentru obținerea ofertelor"""
    try:
        offers = db.query(Offer).order_by(Offer.created_at.desc()).all()
        return [{
            'id': offer.id,
            'offer_id': offer.offer_id,
            'title': _normalize_offer_title(offer.title, offer.offer_id, offer.url),
            'customer': offer.customer,
            'url': offer.url,
            'created_at': offer.created_at.isoformat(),
            'parts_count': len(offer.parts),
            'dosar_id': offer.dosar_id,
            'dosar_path': offer.dosar_path,
            'dosar_allocated': offer.dosar_allocated.isoformat() if offer.dosar_allocated else None,
            'has_dosar': bool(offer.dosar_id),
        } for offer in offers]
    except Exception as e:
        logger.error(f"Eroare la obținerea ofertelor: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/api/offer/{offer_id}/parts")
async def get_offer_parts(offer_id: int, db: Session = Depends(get_db)):
    """API endpoint pentru obținerea reperelor unei oferte"""
    try:
        parts = db.query(Part).filter(Part.offer_id == offer_id).all()
        return [{
            'id': part.id,
            'part_id': part.part_id,
            'part_name': part.name,
            'material': part.material,
            'remarks': part.remarks,
            'weight': part.weight,
            'length': part.length,
            'width': part.width,
            'height': part.height,
            'quantity': part.quantity,
            'unit_price': part.unit_price,
            'discount': part.discount,
            'lead_time': part.lead_time,
            'total_price': part.total_price
        } for part in parts]
    except Exception as e:
        logger.error(f"Eroare la obținerea reperelor: {e}")
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.get("/api/xometry/dosar/{external_offer_id}")
async def get_xometry_dosar_status(
    external_offer_id: str,
    job_id: Optional[str] = None,
    part_ids: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Status dosar pentru extensie: oferta curenta si referinte dupa HJO/J/RFQ sau Part ID."""
    try:
        offer = _find_offer_by_external_id(db, external_offer_id)
        parsed_part_ids = _part_ids_from_raw(part_ids)
        references = _find_dosar_references(db, offer, external_offer_id, job_id, parsed_part_ids)
        references_with_dosar = [item for item in references if item.get("has_dosar")]

        return {
            "success": True,
            "offer_id": external_offer_id,
            "job_id": job_id,
            "part_ids": parsed_part_ids,
            "offer_found": bool(offer),
            "current": _offer_dosar_payload(offer, "current") if offer else None,
            "has_dosar": bool(offer and offer.dosar_id),
            "references": references[:10],
            "references_with_dosar": references_with_dosar[:10],
        }
    except Exception as e:
        logger.error("Could not get Xometry dosar status: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/xometry/dosar/{external_offer_id}/create")
async def create_xometry_dosar(
    external_offer_id: str,
    payload: dict = Body(default_factory=dict),
    db: Session = Depends(get_db),
):
    """Creeaza/aloca dosar pentru oferta Xometry curenta."""
    try:
        offer = _find_offer_by_external_id(db, external_offer_id)
        if not offer:
            title = payload.get("job_name") or payload.get("title") or f"Xometry {external_offer_id}"
            url = payload.get("url") or f"https://partner.xometry.eu/offers/{external_offer_id}"
            offer = Offer(
                offer_id=str(external_offer_id),
                title=title,
                customer="Xometry",
                url=url,
            )
            db.add(offer)
            db.flush()

            for part_data in payload.get("parts") or []:
                part_id = str(part_data.get("part_id") or "").strip()
                if not part_id:
                    continue
                dims = part_data.get("dimensions") or {}
                db.add(Part(
                    offer_id=offer.id,
                    part_id=part_id,
                    name=part_data.get("part_name") or f"Part {part_id}",
                    material=part_data.get("material") or "",
                    quantity=part_data.get("quantity") or 1,
                    length=dims.get("l"),
                    width=dims.get("w"),
                    height=dims.get("h"),
                    processes=part_data.get("process"),
                ))

        if offer.dosar_id:
            return {
                "success": True,
                "message": "Oferta are deja dosar",
                "current": _offer_dosar_payload(offer, "current"),
            }

        from xometry.dosar_service import DosarService

        metadata = {
            "offer_id": str(external_offer_id),
            "job_id": payload.get("job_name") or payload.get("title"),
            "part_ids": [str(item.get("part_id")) for item in payload.get("parts") or [] if item.get("part_id")],
            "url": payload.get("url") or offer.url,
        }
        result = DosarService().allocate_dosar(offer.offer_id, offer.title, metadata=metadata)
        if not result.get("success"):
            raise HTTPException(status_code=502, detail=result.get("error") or "Nu s-a putut crea dosarul")

        offer.dosar_id = result["dosar_id"]
        offer.dosar_path = result.get("path_linux") or result.get("path_windows")
        offer.dosar_allocated = datetime.utcnow()
        db.commit()

        return {
            "success": True,
            "message": "Dosar creat",
            "current": _offer_dosar_payload(offer, "current"),
            "dosar": result,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("Could not create Xometry dosar: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/odoo/dosar/discover")
async def discover_odoo_dosar_action():
    """Incarca action-ul Odoo configurat, util dupa ce pui credentialele in .env."""
    try:
        from xometry.odoo_client import OdooClient

        action = OdooClient().load_dosar_action()
        return {
            "success": True,
            "action_id": action.get("id"),
            "name": action.get("name"),
            "res_model": action.get("res_model"),
            "views": action.get("views"),
            "context": action.get("context"),
        }
    except Exception as e:
        logger.error("Could not discover Odoo dosar action: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/offer/{offer_id}/remarks")
async def update_offer_remarks(offer_id: int, request: Request, db: Session = Depends(get_db)):
    """Actualizează observațiile unei oferte"""
    try:
        data = await request.json()
        remarks = data.get('remarks', '')
        
        offer = db.query(Offer).filter(Offer.id == offer_id).first()
        if not offer:
            raise HTTPException(status_code=404, detail="Oferta nu a fost găsită")
        
        offer.remarks = remarks
        db.commit()
        
        return {"success": True, "message": "Observații ofertă actualizate cu succes"}
        
    except Exception as e:
        logger.error(f"Eroare la actualizarea observațiilor ofertei: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.post("/api/part/{part_id}/remarks")
async def update_part_remarks(part_id: int, request: Request, db: Session = Depends(get_db)):
    """Actualizează observațiile unui reper"""
    try:
        data = await request.json()
        remarks = data.get('remarks', '')
        
        part = db.query(Part).filter(Part.id == part_id).first()
        if not part:
            raise HTTPException(status_code=404, detail="Reperul nu a fost găsit")
        
        part.remarks = remarks
        db.commit()
        
        return {"success": True, "message": "Observații reper actualizate cu succes"}
        
    except Exception as e:
        logger.error(f"Eroare la actualizarea observațiilor reperului: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.post("/api/scrape")
async def scrape_offer_from_extension(request: Request, db: Session = Depends(get_db)):
    """API endpoint pentru primirea datelor de la extensia Chrome"""
    offer_data = None
    try:
        # Parsează datele JSON de la extensie
        offer_data = await request.json()
        offer_external_id = offer_data.get("offer_id") if isinstance(offer_data, dict) else None
        offer_url = offer_data.get("url") if isinstance(offer_data, dict) else None
        normalized_title = _normalize_offer_title(
            offer_data.get("title") if isinstance(offer_data, dict) else None,
            offer_external_id,
            offer_url,
        )
        
        logger.info(f"Primit date de la extensie pentru oferta: {offer_external_id}")

        if not offer_external_id or not offer_url:
            raise HTTPException(status_code=400, detail="Payload invalid: lipsesc offer_id sau url")
        
        # Verifică dacă oferta există deja
        existing_offer = db.query(Offer).filter(Offer.offer_id == offer_external_id).first()
        
        if existing_offer:
            logger.info(f"Oferta {offer_external_id} există deja, actualizez")
            existing_offer.title = normalized_title
            existing_offer.customer = offer_data.get('customer')
            existing_offer.url = offer_url
            offer_id = existing_offer.id
        else:
            # Creează oferta nouă
            offer = Offer(
                offer_id=offer_external_id,
                url=offer_url,
                title=normalized_title,
                customer=offer_data.get('customer')
            )
            db.add(offer)
            db.commit()
            offer_id = offer.id
            logger.info(f"Oferta nouă creată cu ID: {offer_id}")
            
        # Salvează reperele
        parts_added = 0
        current_part_ids = []  # Pentru curățarea imaginilor vechi
        
        for part_data in offer_data.get('parts', []):
            part_id = part_data.get('part_id', f"part_{parts_added}")
            current_part_ids.append(part_id)
            
            # Verifică dacă reperul există
            existing_part = db.query(Part).filter(
                Part.offer_id == offer_id,
                Part.part_id == part_id
            ).first()

            # Procesează procesele - dacă sunt string, convertește la list
            processes = part_data.get("processes", [])
            if isinstance(processes, str):
                processes = [p.strip() for p in processes.split(",") if p.strip()]
            
            # Procesează dimensiunile - extrage din obiectul dimensions
            dimensions = part_data.get("dimensions", {})
            length = dimensions.get("length") if dimensions else part_data.get("length")
            width = dimensions.get("width") if dimensions else part_data.get("width")
            height = dimensions.get("height") if dimensions else part_data.get("height")
            weight = part_data.get("weight")
            
            # Debug logging pentru dimensiuni
            logger.debug(f"Reper {part_id} - Dimensiuni primite: {dimensions}")
            logger.debug(f"Reper {part_id} - L: {length}, W: {width}, H: {height}, Weight: {weight}")
                
            # Procesează imaginea dacă există
            image_url = part_data.get("image_url")
            local_image_path = None
            
            if image_url:
                # Încearcă să descarce imaginea local
                local_image_path = download_and_save_image(
                    image_url, 
                    str(offer_id), 
                    part_id
                )
                
                # Dacă descărcarea a eșuat, păstrează URL-ul extern
                if not local_image_path:
                    logger.warning(f"Nu s-a putut descărca imaginea pentru reperul {part_id}")
                    local_image_path = None

            if existing_part:
                # Actualizează reperul existent
                existing_part.name = part_data.get('part_name', '')
                logger.info(f"Nume nou reper: {existing_part.name}")
                existing_part.material = part_data.get('material', '')
                existing_part.remarks = part_data.get('remarks', '')
                existing_part.weight = weight
                existing_part.length = length
                existing_part.width = width
                existing_part.height = height
                existing_part.quantity = part_data.get('quantity', 1)
                existing_part.unit_price = part_data.get('unit_price')
                existing_part.discount = part_data.get('discount', 0)
                existing_part.lead_time = part_data.get('lead_time')
                existing_part.total_price = part_data.get('total_price')
                existing_part.processes = processes
                if local_image_path:
                    existing_part.local_image_path = local_image_path
                    existing_part.image_url = None # Curăță URL-ul extern dacă avem local
                
            else:
                # Creează reper nou
                new_part = Part(
                    offer_id=offer_id,
                    part_id=part_id,
                    name=part_data.get('part_name', ''),
                    material=part_data.get('material', ''),
                    remarks=part_data.get('remarks', ''),
                    weight=weight,
                    length=length,
                    width=width,
                    height=height,
                    quantity=part_data.get('quantity', 1),
                    unit_price=part_data.get('unit_price'),
                    discount=part_data.get('discount', 0),
                    lead_time=part_data.get('lead_time'),
                    total_price=part_data.get('total_price'),
                    image_url=image_url if not local_image_path else None,
                    local_image_path=local_image_path,
                    processes=processes
                )
                db.add(new_part)
                parts_added += 1

        db.commit()
        
        # Opțional: Curățare imagini vechi pentru reperele care nu mai există
        # (Codul existent pentru curățare poate fi apelat aici dacă e necesar)

        return {"success": True, "message": f"Ofertă salvată. {parts_added} repere adăugate/actualizate."}

    except HTTPException:
        db.rollback()
        raise
    except Exception:
        payload_keys = sorted(offer_data.keys()) if isinstance(offer_data, dict) else []
        logger.exception("Eroare la salvarea ofertei. Payload keys=%s", payload_keys)
        db.rollback()
        raise HTTPException(status_code=500, detail="Eroare internă")

@app.post("/api/orders/sync", response_model=Dict[str, str], tags=["Orders"])
async def sync_orders(payload: OrderSyncRequest, db: Session = Depends(get_db)):
    """
    Sincronizează lista de comenzi din extensie în baza de date.
    Descarcă și salvează imaginile local.
    """
    try:
        orders = payload.orders
        logger.info(f"Received {len(orders)} orders for sync.")

        synced_count = 0
        import re
        def _strip_ext(name: str) -> str:
            if not name:
                return name
            name = name.strip()
            m = re.match(r"^(.*)\.([A-Za-z0-9]{1,10})$", name)
            if not m:
                return name
            return m.group(1) or name

        for order_data in orders:
            od = order_data.dict()
            order_id = od.get('order_id')
            part_id = od.get('part_id')
            
            if not order_id:
                continue

            # Normalize part name (remove file extension)
            if od.get('part_name'):
                od['part_name'] = _strip_ext(od.get('part_name'))
            if od.get('part_name') is None:
                od['part_name'] = None

            # Descarcă imaginea local dacă există un URL sau date Base64
            img_url = od.get('Image') or od.get('image_url')
            img_data = od.get('image_data')
            local_img_path = None
            
            if img_data:
                try:
                    local_img_path = save_base64_image(img_data, f"orders/{order_id}", part_id or 'part')
                    if local_img_path:
                        logger.info(f"✅ Saved base64 image for {order_id}/{part_id}")
                    else:
                        logger.warning(f"❌ Failed to save base64 image for {order_id}/{part_id}")
                except Exception as b64_err:
                    logger.warning(f"Failed to save base64 image: {b64_err}")
            
            if not local_img_path and img_url and img_url.startswith('http'):
                try:
                    # Folosim order_id ca folder pentru organizare
                    local_img_path = download_and_save_image(img_url, f"orders/{order_id}", part_id or 'part')
                    if local_img_path:
                        logger.info(f"✅ Downloaded image for {order_id}/{part_id}")
                except Exception as img_err:
                    logger.warning(f"Failed to download image: {img_err}")

            # Check if order exists
            query = db.query(Order).filter(Order.order_id == order_id)
            if part_id:
                query = query.filter(Order.part_id == part_id)
            
            existing = query.first()

            if existing:
                existing.status = od.get('status')
                existing.price = od.get('price')
                if local_img_path:
                    existing.local_image_path = local_img_path
                existing.updated_at = datetime.utcnow()
                existing.details = od
            else:
                new_order = Order(
                    order_id=order_id,
                    part_id=part_id,
                    status=od.get('status'),
                    order_date=od.get('date'),
                    price=od.get('price'),
                    local_image_path=local_img_path,
                    details=od
                )
                db.add(new_order)
                synced_count += 1
        
        db.commit()
        return {"success": "True", "message": f"Synced {synced_count} orders."}

    except Exception as e:
        logger.error(f"Error syncing orders: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/orders/check/{part_id}", response_model=OrderCheckResponse, tags=["Orders"])
async def check_order_history(part_id: str, db: Session = Depends(get_db)):
    """
    Verifică istoricul comenzilor pentru un Part ID specific.
    Returnează detaliile ultimei comenzi dacă există.
    """
    try:
        # Search by Part ID
        order = db.query(Order).filter(Order.part_id == part_id).order_by(Order.created_at.desc()).first()
        
        if order:
            return {
                "exists": True,
                "order_id": order.order_id,
                "date": order.order_date,
                "status": order.status,
                "price": order.price
            }
        
        return {"exists": False}

    except Exception as e:
        logger.error(f"Error checking order history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/allocate-dosar/{offer_id}")
async def allocate_dosar(offer_id: int, db: Session = Depends(get_db)):
    """Alocă un dosar pentru o ofertă"""
    try:
        # Obține oferta
        offer = db.query(Offer).filter(Offer.id == offer_id).first()
        if not offer:
            raise HTTPException(status_code=404, detail="Oferta nu a fost găsită")
        
        # Verifică dacă oferta are deja un dosar alocat
        if offer.dosar_id:
            return {
                "success": True,
                "message": "Oferta are deja un dosar alocat",
                "dosar_id": offer.dosar_id,
                "dosar_path": offer.dosar_path,
                "allocated_at": offer.dosar_allocated.isoformat() if offer.dosar_allocated else None
            }
        
        # Importă serviciul de dosare
        from xometry.dosar_service import DosarService
        dosar_service = DosarService()
        
        # Alocă dosarul
        result = dosar_service.allocate_dosar(offer.offer_id, offer.title)
        
        if result['success']:
            # Actualizează oferta cu informațiile dosarului
            offer.dosar_id = result['dosar_id']
            offer.dosar_path = result['path_linux']
            offer.dosar_allocated = datetime.utcnow()
            
            db.commit()
            
            logger.info(f"Dosar alocat pentru oferta {offer_id}: {result['dosar_id']}")
            
            return {
                "success": True,
                "message": "Dosar alocat cu succes",
                "dosar_id": result['dosar_id'],
                "folder_name": result['folder_name'],
                "path_linux": result['path_linux'],
                "path_windows": result['path_windows'],
                "allocated_at": result['allocated_at']
            }
        else:
            logger.error(f"Eroare la alocarea dosarului pentru oferta {offer_id}: {result.get('error')}")
            raise HTTPException(status_code=500, detail=f"Eroare la alocarea dosarului: {result.get('error')}")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Eroare la alocarea dosarului: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Eroare internă: {str(e)}")

@app.get("/api/offer/{offer_id}/dosar")
async def get_dosar_info(offer_id: int, db: Session = Depends(get_db)):
    """Obține informațiile despre dosarul unei oferte"""
    try:
        offer = db.query(Offer).filter(Offer.id == offer_id).first()
        if not offer:
            raise HTTPException(status_code=404, detail="Oferta nu a fost găsită")
        
        if not offer.dosar_id:
            return {
                "success": True,
                "has_dosar": False,
                "message": "Oferta nu are dosar alocat"
            }
        
        return {
            "success": True,
            "has_dosar": True,
            "dosar_id": offer.dosar_id,
            "dosar_path": offer.dosar_path,
            "allocated_at": offer.dosar_allocated.isoformat() if offer.dosar_allocated else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Eroare la obținerea informațiilor despre dosar: {e}")
        raise HTTPException(status_code=500, detail=f"Eroare internă: {str(e)}")


@app.post("/api/extension/analyze")
async def analyze_part_extension_deprecated(
    payload: dict = Body(...),
    db: Session = Depends(get_db)
):
    """
    Endpoint pentru extensia Chrome.
    Primește part_id și returnează date despre reper (grosime, îndoiri, preț).
    """
    try:
        part_id = payload.get("part_id")
        if not part_id:
            raise HTTPException(status_code=400, detail="part_id is required")

        # Caută reperul (exact sau partial)
        part = db.query(Part).filter(Part.part_id == part_id).first()
        if not part:
            # Încercăm căutare parțială
            part = db.query(Part).filter(Part.part_id.like(f"%{part_id}%")).first()

        if not part:
            return {"found": False, "message": "Part not found"}

        # Logică pentru extragere date
        thickness = None
        bends = 0

        # Grosime din dimensiuni (Height)
        if part.height and part.height > 0:
             thickness = part.height
        
        # Grosime din Remarks (fallback)
        if not thickness and part.remarks:
            match = re.search(r"(?:Gauge|Thickness):\s*([\d\.]+)\s*mm", part.remarks, re.IGNORECASE)
            if match:
                try:
                    thickness = float(match.group(1))
                except:
                    pass

        return {
            "found": True,
            "part_id": part.part_id,
            "data": {
                "observations": part.remarks,
                "thickness": thickness,
                "bends": bends,
                "price": part.unit_price,
                "currency": "EUR"
            }
        }

    except Exception as e:
        logger.error(f"Error in extension analysis: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.websocket("/ws/{part_id}")
async def websocket_endpoint(websocket: WebSocket, part_id: str):
    await manager.connect(websocket, part_id)
    try:
        while True:
            # Keep connection alive, maybe wait for client messages if needed
            data = await websocket.receive_text()
            # Echo or ignore
            pass
    except WebSocketDisconnect:
        manager.disconnect(websocket, part_id)
    except Exception as e:
        logger.error(f"WS Error: {e}")
        manager.disconnect(websocket, part_id)


async def simulate_analysis_task(part_id: str, db_session):
    """
    Simulates long running analysis (20s) then notifies WS.
    Note: Creating a new DB session here manually would be better, 
    but for simulation we might re-query or pass data needed.
    """
    logger.info(f"Starting analysis for {part_id} (20s delay)")
    await asyncio.sleep(20) # Simulate delay
    
    # Re-query Logic (need to handle DB session safely in thread/async)
    # Since db is passed from route, it might close? 
    # FastAPI Depends(get_db) yields, so it closes after request.
    # We should act on data we extracted or create a new session.
    # For now, let's extract data inside the route and pass it to this task.
    
    # Actually, simpler: just pass the payload we want to send back.
    # But we want to simulate "processing". 
    pass

async def process_and_notify(part_id: str, data: dict):
    await asyncio.sleep(20)
    msg = {
        "type": "analysis_complete",
        "part_id": part_id,
        "data": data
    }
    logger.info(f"Broadcasting update for {part_id}")
    await manager.broadcast_to_part(msg, part_id)


@app.post("/api/extension/analyze")
async def analyze_part_extension(
    background_tasks: BackgroundTasks,
    payload: dict = Body(...),
    db: Session = Depends(get_db)
):
    """
    Async endpoint with WebSocket notification.
    """
    try:
        part_id = payload.get("part_id")
        if not part_id:
            raise HTTPException(status_code=400, detail="part_id is required")

        # Basic lookup sync
        part = db.query(Part).filter(Part.part_id == part_id).first()
        if not part:
             part = db.query(Part).filter(Part.part_id.like(f"%{part_id}%")).first()

        if not part:
            return {"found": False, "message": "Part not found"}

        # Extract Data
        thickness = None
        bends = 0
        if part.height and part.height > 0:
             thickness = part.height
        
        if not thickness and part.remarks:
            match = re.search(r"(?:Gauge|Thickness):\s*([\d\.]+)\s*mm", part.remarks, re.IGNORECASE)
            if match:
                try:
                    thickness = float(match.group(1))
                except:
                    pass

        result_data = {
            "observations": part.remarks,
            "thickness": thickness,
            "bends": bends,
            "price": part.unit_price,
            "currency": "EUR"
        }

        # Queue background task
        background_tasks.add_task(process_and_notify, part.part_id, result_data)

        # Immediate return
        return {
            "status": "processing",
            "message": "Analysis started. Wait for Websocket notification.",
            "part_id": part.part_id
        }

    except Exception as e:
        logger.error(f"Error in async analysis: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/orders", tags=["Orders"])
async def get_orders_api(
    q: Optional[str] = None, 
    limit: int = 50, 
    offset: int = 0, 
    db: Session = Depends(get_db)
):
    """API for dynamic order loading and searching"""
    try:
        query = db.query(Order)
        if q:
            term = f"%{q.lower()}%"
            # Search in common fields. Searching in JSON cast to string is a bit slow but works for SQLite/Postgres.
            query = query.filter(
                (Order.order_id.ilike(term)) | 
                (Order.part_id.ilike(term)) |
                (Order.status.ilike(term)) |
                (Order.details.cast(db_String).ilike(term))
            )
        
        total = query.count()
        
        # We need to sort by business date (DD.MM.YYYY string). 
        # In SQLite: substr(date, 7, 4) || substr(date, 4, 2) || substr(date, 1, 2) for YYYYMMDD
        from sqlalchemy import func
        sortable_date = func.substr(Order.order_date, 7, 4).op('||')(func.substr(Order.order_date, 4, 2)).op('||')(func.substr(Order.order_date, 1, 2))
        
        # Sort by that business date DESC, then by internal ID DESC
        orders = query.order_by(sortable_date.desc(), Order.id.desc()).offset(offset).limit(limit).all()
        
        return {
            "total": total,
            "offset": offset,
            "limit": limit,
            "orders": [
                {
                    "id": o.id,
                    "order_id": o.order_id,
                    "part_id": o.part_id,
                    "status": o.status,
                    "order_date": o.order_date,
                    "price": o.price,
                    "local_image": o.local_image_path,
                    "details": o.details
                } for o in orders
            ]
        }
    except Exception as e:
        logger.error(f"Error in get_orders_api: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/orders", response_class=HTMLResponse)
async def list_orders(request: Request):
    """Afișează dashboard-ul (initial page only)"""
    return templates.TemplateResponse("orders.html", {"request": request})

@app.get("/documentation", response_class=HTMLResponse)
async def documentation(request: Request):
    """Pagina de documentație"""
    return templates.TemplateResponse("documentation.html", {"request": request})

@app.get("/help", response_class=HTMLResponse)
async def help_page(request: Request):
    """Pagina de ajutor"""
    return templates.TemplateResponse("help.html", {"request": request})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=10000)
